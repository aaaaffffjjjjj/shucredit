"""分析悉尼工商学院Excel中颜色与模块的对应关系"""
import os
import openpyxl

excel_path = os.path.join(os.path.dirname(__file__), 'shucredit_scripts', 'pdf_app', 'uploads', 'sydney_business_25.xlsx')

print("="*100)
print("分析颜色与模块的对应关系")
print("="*100)

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['信管课程25级']

# 收集课程-模块-颜色关系
course_info = {}

for row_idx in range(1, sheet.max_row + 1):
    row = sheet[row_idx]
    
    # 获取这一行的模块信息（右侧列）
    row_modules = []
    for col_idx in range(6, sheet.max_column + 1):  # 从第6列开始是模块信息
        if col_idx - 1 >= len(row):
            continue
        cell = row[col_idx - 1]
        if cell.value is None or str(cell.value).strip() == "":
            continue
        cell_content = str(cell.value).strip()
        
        # 过滤无效内容
        invalid_phrases = ["未必能完全", "注意", "推测", "简称建议", "选双学位", "各总计"]
        is_invalid = any(phrase in cell_content for phrase in invalid_phrases)
        if is_invalid:
            continue
        
        # 清理模块名
        module_name = cell_content
        module_name = module_name.replace('（', '(').replace('）', ')')
        module_name = module_name.split('(')[0].strip()
        if module_name:
            row_modules.append(module_name)
    
    # 获取这一行的课程及其颜色
    for col_idx in range(1, 6):  # 前5列是课程列
        if col_idx - 1 >= len(row):
            continue
        cell = row[col_idx - 1]
        if cell.value is None or str(cell.value).strip() == "":
            continue
        
        course_name = str(cell.value).strip()
        # 过滤无效值
        invalid_phrases = ["大一", "大二", "大三", "大四", "前半学期", "后半学期", "夏"]
        is_invalid = any(phrase in course_name for phrase in invalid_phrases)
        if is_invalid or len(course_name) < 2:
            continue
        
        # 获取颜色信息
        font_color = None
        fill_color = None
        
        if cell.font and cell.font.color and cell.font.color.rgb:
            font_color = cell.font.color.rgb
        
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            fill_color = cell.fill.fgColor.rgb
        
        # 保存信息
        if course_name not in course_info:
            course_info[course_name] = {
                'font_colors': set(),
                'fill_colors': set(),
                'modules': set()
            }
        
        if font_color:
            course_info[course_name]['font_colors'].add(font_color)
        if fill_color:
            course_info[course_name]['fill_colors'].add(fill_color)
        for module in row_modules:
            course_info[course_name]['modules'].add(module)

# 按填充颜色分组
fill_color_groups = {}
for course_name, info in course_info.items():
    for color in info['fill_colors']:
        if color not in fill_color_groups:
            fill_color_groups[color] = {
                'courses': [],
                'modules': set()
            }
        fill_color_groups[color]['courses'].append(course_name)
        for module in info['modules']:
            fill_color_groups[color]['modules'].add(module)

print("\n按填充颜色分组:")
print("="*100)
for color, group in fill_color_groups.items():
    print(f"\n填充颜色: {color}")
    print(f"  课程数: {len(group['courses'])}")
    print(f"  关联模块: {', '.join(group['modules']) if group['modules'] else '-'}")
    print(f"  课程示例: {', '.join(group['courses'][:5])}")

# 按字体颜色分组
font_color_groups = {}
for course_name, info in course_info.items():
    for color in info['font_colors']:
        if color not in font_color_groups:
            font_color_groups[color] = {
                'courses': [],
                'modules': set()
            }
        font_color_groups[color]['courses'].append(course_name)
        for module in info['modules']:
            font_color_groups[color]['modules'].add(module)

print("\n\n按字体颜色分组:")
print("="*100)
for color, group in font_color_groups.items():
    print(f"\n字体颜色: {color}")
    print(f"  课程数: {len(group['courses'])}")
    print(f"  关联模块: {', '.join(group['modules']) if group['modules'] else '-'}")
    print(f"  课程示例: {', '.join(group['courses'][:5])}")
