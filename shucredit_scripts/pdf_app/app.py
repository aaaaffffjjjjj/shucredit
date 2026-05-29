#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
上海大学培养方案PDF解析器 - Web应用版
功能：上传PDF → 智能解析 → 预览编辑 → 导入数据库
新增功能：通过PDF文件名识别专业和学院归属
"""

import os
import re
import pdfplumber
import pymysql
import uuid
import pandas as pd
from flask import Flask, request, render_template, jsonify, redirect, url_for

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['SECRET_KEY'] = 'supersecretkey'

# 确保目录存在
os.makedirs('uploads', exist_ok=True)
os.makedirs('templates', exist_ok=True)

# 数据库配置
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '24567@Zzy',
    'database': 'student_system',
    'charset': 'utf8mb4'
}

# 全局状态（用于存储解析结果）
parsing_results = {}

# ==================== 学院-专业映射表 ====================
COLLEGE_MAJOR_MAP = {
    # 通信与信息工程学院
    '通信与信息工程学院': [
        '通信工程', '电子信息工程', '电子科学与技术', '信息安全',
        '通信工程(中外合作)', '电子信息工程(中外合作)'
    ],
    # 微电子学院
    '微电子学院': [
        '微电子科学与工程', '集成电路设计与集成系统', '微电子科学与工程(中外合作)'
    ],
    # 计算机工程与科学学院
    '计算机工程与科学学院': [
        '计算机科学与技术', '软件工程', '网络工程', '人工智能',
        '计算机科学与技术(中外合作)', '数据科学与大数据技术'
    ],
    # 机电工程与自动化学院
    '机电工程与自动化学院': [
        '机械工程', '机械设计制造及其自动化', '自动化', '测控技术与仪器',
        '工业工程', '机器人工程', '智能制造工程'
    ],
    # 材料科学与工程学院
    '材料科学与工程学院': [
        '材料科学与工程', '材料物理', '材料化学', '新能源材料与器件'
    ],
    # 环境与化学工程学院
    '环境与化学工程学院': [
        '环境工程', '化学工程与工艺', '应用化学', '生物工程', '制药工程'
    ],
    # 生命科学学院
    '生命科学学院': [
        '生物科学', '生物技术', '生物信息学', '食品科学与工程'
    ],
    # 理学院
    '理学院': [
        '数学与应用数学', '信息与计算科学', '物理学', '应用物理学',
        '统计学', '光电信息科学与工程'
    ],
    # 外国语学院
    '外国语学院': [
        '英语', '日语', '德语', '法语', '翻译'
    ],
    # 文学院
    '文学院': [
        '汉语言文学', '历史学', '哲学', '新闻学', '广播电视学', '广告学'
    ],
    # 法学院
    '法学院': [
        '法学', '知识产权'
    ],
    # 经济学院
    '经济学院': [
        '经济学', '国际经济与贸易', '金融学', '会计学', '工商管理',
        '市场营销', '信息管理与信息系统', '工程管理'
    ],
    # 管理学院
    '管理学院': [
        '工商管理', '公共管理', '信息管理与信息系统', '物流管理', '工业工程'
    ],
    # 悉尼工商学院
    '悉尼工商学院': [
        '国际经济与贸易', '工商管理', '会计学', '金融学', '信息管理与信息系统'
    ],
    # 中欧工程技术学院
    '中欧工程技术学院': [
        '机械工程', '电气工程及其自动化', '材料科学与工程', '生物工程'
    ],
    # 社会学院
    '社会学院': [
        '社会学', '社会工作', '公共事业管理'
    ],
    # 美术学院
    '美术学院': [
        '美术学', '绘画', '雕塑', '视觉传达设计', '环境设计', '产品设计'
    ],
    # 音乐学院
    '音乐学院': [
        '音乐表演', '音乐学', '作曲与作曲技术理论'
    ],
    # 体育学院
    '体育学院': [
        '体育教育', '运动训练', '社会体育指导与管理'
    ],
    # 影视学院
    '影视学院': [
        '戏剧影视文学', '广播电视编导', '表演', '动画', '数字媒体艺术'
    ],
    # 新闻传播学院
    '新闻传播学院': [
        '新闻学', '广播电视学', '广告学', '网络与新媒体', '传播学'
    ],
    # 马克思主义学院
    '马克思主义学院': [
        '思想政治教育'
    ],
    # 力学与工程科学学院
    '力学与工程科学学院': [
        '工程力学', '土木工程', '建筑环境与能源应用工程'
    ],
    # 土木工程系
    '土木工程系': [
        '土木工程', '建筑环境与能源应用工程', '给排水科学与工程'
    ],
    # 电气工程系
    '电气工程系': [
        '电气工程及其自动化', '自动化', '智能电网信息工程'
    ],
    # 化学系
    '化学系': [
        '化学', '应用化学', '材料化学'
    ],
    # 物理系
    '物理系': [
        '物理学', '应用物理学', '光电信息科学与工程'
    ],
    # 数学系
    '数学系': [
        '数学与应用数学', '信息与计算科学', '统计学'
    ],
    # 工程训练中心
    '工程训练中心': [
        '智能制造工程', '工业工程'
    ]
}

# 反向映射：专业名称 → 学院名称
MAJOR_COLLEGE_MAP = {}
for college, majors in COLLEGE_MAJOR_MAP.items():
    for major in majors:
        MAJOR_COLLEGE_MAP[major] = college

# 常见专业名称关键词
MAJOR_KEYWORDS = [
    '通信工程', '电子信息工程', '微电子', '计算机', '软件工程',
    '机械', '自动化', '材料', '环境', '化学', '生物',
    '数学', '物理', '英语', '日语', '法学', '经济', '金融',
    '会计', '管理', '土木', '电气', '建筑', '设计', '音乐',
    '体育', '新闻', '传播', '影视', '美术', '哲学', '历史'
]


def guess_college_and_major_from_filename(filename):
    """
    从PDF文件名智能识别学院和专业
    
    Args:
        filename: PDF文件名
        
    Returns:
        (college_name, major_name): 学院名称和专业名称
    """
    # 移除扩展名和常见前缀
    name = filename.replace('.pdf', '').replace('.PDF', '')
    name = re.sub(r'^(202[0-9]|培养方案|本科|专业)', '', name)
    
    # 查找专业关键词
    found_major = None
    for keyword in MAJOR_KEYWORDS:
        if keyword in name:
            found_major = keyword
            break
    
    # 如果找到专业关键词，尝试匹配学院
    if found_major:
        # 精确匹配
        for major, college in MAJOR_COLLEGE_MAP.items():
            if found_major in major:
                return college, major
        
        # 模糊匹配
        for major, college in MAJOR_COLLEGE_MAP.items():
            if found_major in major or major in found_major:
                return college, major
        
        return "未知学院", found_major
    
    return "未知学院", "未知专业"


def get_all_colleges():
    """获取所有学院列表"""
    return list(COLLEGE_MAJOR_MAP.keys())


def get_majors_by_college(college_name):
    """根据学院获取专业列表"""
    return COLLEGE_MAJOR_MAP.get(college_name, [])


# ==================== SmartPDFParser类 ====================
class SmartPDFParser:
    """智能PDF解析器"""
    
    def __init__(self):
        self.modules = []
        self.courses = []
        self.full_text = ""
        self.college_name = ""
        self.major_name = ""
        self.filename = ""
    
    def parse(self, pdf_path, filename=""):
        """解析PDF"""
        self.filename = filename
        
        # 1. 先从文件名识别学院和专业
        if filename:
            college_from_file, major_from_file = guess_college_and_major_from_filename(filename)
            self.college_name = college_from_file
            self.major_name = major_from_file
        
        # 2. 解析PDF内容
        with pdfplumber.open(pdf_path) as pdf:
            self.full_text = "\n".join([page.extract_text() or "" for page in pdf.pages])
        
        # 3. 从PDF内容提取学院和专业（如果文件名没识别出来）
        self._extract_college_and_major()
        
        # 4. 解析模块（从PDF提取）
        self._parse_modules_from_pdf()
        
        # 5. 解析课程（从PDF提取）
        self._parse_courses_from_pdf()
        
        return True
    
    def _extract_college_and_major(self):
        """从PDF内容提取学院和专业名称"""
        # 如果文件名已经识别出学院和专业，优先级较低
        patterns = [
            r'(\S+学院)\s+(\S+专业)',
            r'(\S+专业)\s+培养方案',
            r'(\S+学院)',
            r'(\S+专业)'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, self.full_text)
            if match:
                if len(match.groups()) >= 2:
                    pdf_college = match.group(1)
                    pdf_major = match.group(2)
                else:
                    pdf_major = match.group(1)
                    pdf_college = None
                
                # 更新学院和专业（优先使用PDF内容）
                if pdf_college and pdf_college != "未知学院":
                    self.college_name = pdf_college
                if pdf_major and pdf_major != "未知专业":
                    self.major_name = pdf_major
                break
    
    def _parse_modules_from_pdf(self):
        """从PDF解析模块，按照数字编号构建层级结构"""
        lines = self.full_text.split('\n')
        modules = []
        current_root_order = 0
        
        for line in lines:
            # 匹配根模块：1. 模块名 学分 或 1、模块名 学分
            root_match = re.match(r'^(\d+)[\.、]\s*([^\n]+?)\s*(\d+(?:\.\d+)?)\s*学分', line.strip())
            if root_match:
                current_root_order = int(root_match.group(1))
                modules.append({
                    'order': current_root_order,
                    'sub_order': None,  # 根模块
                    'name': root_match.group(2).strip(),
                    'credits': float(root_match.group(3)),
                    'parent_order': None
                })
                continue
            
            # 匹配子模块：（1）模块名 学分 或 (1) 模块名 学分
            sub_match = re.match(r'^[（(](\d+)[）)]\s*([^\n]+?)\s*(\d+(?:\.\d+)?)\s*学分', line.strip())
            if sub_match and current_root_order > 0:
                modules.append({
                    'order': current_root_order,
                    'sub_order': int(sub_match.group(1)),
                    'name': sub_match.group(2).strip(),
                    'credits': float(sub_match.group(3)),
                    'parent_order': current_root_order
                })
                continue
        
        # 去重
        seen = set()
        unique_modules = []
        for mod in modules:
            key = (mod['order'], mod.get('sub_order'), mod['name'], mod['credits'])
            if key not in seen:
                seen.add(key)
                unique_modules.append(mod)
        
        # 排序：先按主序号，再按子序号（None值排在前面）
        unique_modules.sort(key=lambda x: (x['order'], x.get('sub_order') is None, x.get('sub_order') or 0))
        self._map_to_hierarchy_structure(unique_modules)
    
    def _map_to_hierarchy_structure(self, parsed_modules):
        """将解析的模块按照层级结构组织"""
        modules = []
        module_id = 1
        order_id_map = {}  # order -> id 映射
        
        # 先创建所有根模块
        root_modules = [m for m in parsed_modules if m['sub_order'] is None]
        for mod in root_modules:
            modules.append({
                'id': module_id,
                'name': mod['name'],
                'required_credits': mod['credits'],
                'parent_id': None,
                'source': 'pdf'
            })
            order_id_map[mod['order']] = module_id
            module_id += 1
        
        # 然后创建所有子模块
        sub_modules = [m for m in parsed_modules if m['sub_order'] is not None]
        for mod in sub_modules:
            parent_id = order_id_map.get(mod['parent_order'])
            if parent_id:
                modules.append({
                    'id': module_id,
                    'name': mod['name'],
                    'required_credits': mod['credits'],
                    'parent_id': parent_id,
                    'source': 'pdf'
                })
                module_id += 1
        
        self.modules = modules
    
    def _map_to_standard_structure(self, parsed_modules):
        """将解析的模块映射到标准结构（旧方法，保留兼容）"""
        standard_names = [
            '公共基础课程', '通识课程', '专业基础课程', 
            '专业必修课程', '专业选修课程', '个性化教育课程', '专业方向模块'
        ]
        
        std_modules = []
        module_id = 1
        
        for name in standard_names:
            found = False
            for parsed in parsed_modules:
                if name in parsed['name'] or parsed['name'] in name:
                    std_modules.append({
                        'id': module_id,
                        'name': name,
                        'required_credits': parsed['credits'],
                        'parent_id': None,
                        'source': 'pdf'
                    })
                    module_id += 1
                    found = True
                    break
            
            if not found:
                default_credits = {
                    '公共基础课程': 67.5,
                    '通识课程': 8.0,
                    '专业基础课程': 52.0,
                    '专业必修课程': 18.5,
                    '专业选修课程': 4.0,
                    '个性化教育课程': 10.0,
                    '专业方向模块': 10.0
                }
                std_modules.append({
                    'id': module_id,
                    'name': name,
                    'required_credits': default_credits.get(name, 10.0),
                    'parent_id': None,
                    'source': 'default'
                })
                module_id += 1
        
        sub_modules = [
            ('思政类', 18.5, '公共基础课程'),
            ('军体类', 9.0, '公共基础课程'),
            ('大学英语', 8.0, '公共基础课程'),
            ('人工智能类', 5.0, '公共基础课程'),
            ('国家安全教育', 1.0, '公共基础课程'),
            ('自然科学类', 26.0, '公共基础课程'),
            ('核心通识课', 2.0, '通识课程'),
            ('跨类通识课', 2.0, '通识课程'),
            ('其他通识课', 4.0, '通识课程'),
            ('专业选修子模块1', 2.0, '专业选修课程'),
            ('专业选修子模块2', 2.0, '专业选修课程'),
            ('综合工程实践能力培养模块', 10.0, '个性化教育课程'),
        ]
        
        direction_credits = self._parse_direction_credits()
        if direction_credits:
            sub_modules.extend(direction_credits)
        else:
            sub_modules.extend([
                ('方向1', 4.0, '专业方向模块'),
                ('方向2', 4.0, '专业方向模块'),
                ('方向3', 2.0, '专业方向模块'),
            ])
        
        for name, credits, parent_name in sub_modules:
            parent_id = next((m['id'] for m in std_modules if m['name'] == parent_name), None)
            if parent_id:
                std_modules.append({
                    'id': module_id,
                    'name': name,
                    'required_credits': credits,
                    'parent_id': parent_id,
                    'source': 'default'
                })
                module_id += 1
        
        self.modules = std_modules
    
    def _parse_direction_credits(self):
        """尝试从PDF解析专业方向"""
        direction_section = re.search(r'专业方向.*?(?=\n\n|\Z)', self.full_text, re.DOTALL)
        if direction_section:
            text = direction_section.group()
            pattern = r'(\S+方向)\s*(\d+(?:\.\d+)?)学分'
            matches = re.findall(pattern, text)
            if matches:
                return [(m[0], float(m[1]), '专业方向模块') for m in matches]
        return None
    
    def _parse_courses_from_pdf(self):
        """从PDF解析课程"""
        courses = []
        lines = self.full_text.split('\n')
        course_id = 100
        
        # 创建模块名到ID的映射
        module_name_to_id = {}
        for mod in self.modules:
            module_name_to_id[mod['name']] = mod['id']
        
        for i, line in enumerate(lines):
            code_match = re.match(r'([A-Z]{3}\d{5,7})', line.strip())
            if code_match:
                code = code_match.group(1)
                name = ""
                credit = 2.0
                
                # 向上查找最近的模块，作为课程所属模块
                module_ids = []
                module_names = []
                for j in range(i - 1, max(i - 50, 0), -1):
                    prev_line = lines[j].strip()
                    # 检查这一行是否是模块
                    root_match = re.match(r'^(\d+)[\.、]\s*([^\n]+?)\s*(\d+(?:\.\d+)?)\s*学分', prev_line)
                    sub_match = re.match(r'^[（(](\d+)[）)]\s*([^\n]+?)\s*(\d+(?:\.\d+)?)\s*学分', prev_line)
                    
                    if root_match or sub_match:
                        module_name = root_match.group(2).strip() if root_match else sub_match.group(2).strip()
                        # 匹配模块名称
                        for known_name, known_id in module_name_to_id.items():
                            if module_name in known_name or known_name in module_name:
                                if known_id not in module_ids:
                                    module_ids.append(known_id)
                                    module_names.append(known_name)
                        if module_ids:
                            break
                
                for j in range(i + 1, min(i + 5, len(lines))):
                    next_line = lines[j].strip()
                    if next_line and not re.match(r'[A-Z]{3}\d+', next_line):
                        credit_match = re.search(r'(\d+(?:\.\d+)?)\s*学分', next_line)
                        if credit_match:
                            credit = float(credit_match.group(1))
                            name = next_line[:next_line.find(credit_match.group(0))].strip()
                        else:
                            name = next_line
                        break
                
                if name:
                    # 如果没有找到模块，使用默认值
                    if not module_ids:
                        module_ids = [3]
                        # 查找模块3的名称
                        for known_name, known_id in module_name_to_id.items():
                            if known_id == 3:
                                module_names = [known_name]
                                break
                    
                    primary_module_id = module_ids[0] if module_ids else 1
                    
                    courses.append({
                        'id': course_id,
                        'course_code': code,
                        'name': name,
                        'credits': credit,
                        'module_id': primary_module_id,
                        'module_ids': module_ids,
                        'module_names': module_names
                    })
                    course_id += 1
        
        self.courses = courses
    
    def get_result(self):
        """获取解析结果"""
        return {
            'college_name': self.college_name,
            'major_name': self.major_name,
            'filename': self.filename,
            'modules': self.modules,
            'courses': self.courses
        }


# ==================== SmartExcelParser类 ====================
class SmartExcelParser:
    """智能Excel解析器"""

    def __init__(self):
        self.modules = []
        self.courses = []
        self.college_name = ""
        self.major_name = ""
        self.filename = ""
        self.sheet_name = ""

    def parse(self, excel_path, sheet_name, filename=""):
        """解析Excel指定工作表"""
        self.filename = filename
        self.sheet_name = sheet_name

        # 1. 先从文件名识别学院和专业
        if filename:
            college_from_file, major_from_file = guess_college_and_major_from_filename(filename)
            self.college_name = college_from_file
            self.major_name = major_from_file

        # 2. 从工作表名识别专业
        if sheet_name and not self.major_name:
            self._extract_major_from_sheet_name(sheet_name)

        # 3. 解析Excel内容
        df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

        # 4. 检测Excel格式
        is_sydney_format = self._detect_sydney_format(df)

        # 5. 根据格式选择解析策略
        if is_sydney_format:
            # 如果是悉尼工商学院格式，先尝试从同一个文件获取学分概况信息
            try:
                excel_file = pd.ExcelFile(excel_path)
                # 找对应的学分食用概况表
                credit_sheet_name = None
                for name in excel_file.sheet_names:
                    if "学分食用概况" in name:
                        # 尝试匹配专业
                        if sheet_name[:2] in name:  # 简单匹配，如"信管"匹配"信管25级学分食用概况"
                            credit_sheet_name = name
                            break
                        if not credit_sheet_name:  # 如果没有匹配的，取第一个
                            credit_sheet_name = name
                
                credit_df = None
                if credit_sheet_name:
                    credit_df = pd.read_excel(excel_path, sheet_name=credit_sheet_name, header=None)
                
                # 解析悉尼工商学院格式（使用颜色匹配）
                self._parse_sydney_format_with_color(excel_path, sheet_name, credit_df)
            except Exception as e:
                print(f"解析悉尼格式时出错: {e}")
                # 如果出错，回退到标准格式
                self._parse_modules_from_excel(df)
                self._parse_courses_from_excel(df)
        else:
            # 标准格式解析
            self._parse_modules_from_excel(df)
            self._parse_courses_from_excel(df)

        return True
    
    def _detect_sydney_format(self, df):
        """检测是否是悉尼工商学院格式"""
        # 悉尼工商学院格式特点：
        # 1. 有"大一上"、"大一下"等学期列
        # 2. 右侧有模块信息列
        has_semester = False
        has_module_info = False
        
        for i in range(min(10, len(df))):
            row = df.iloc[i].values
            for j, cell in enumerate(row):
                if pd.notna(cell):
                    cell_str = str(cell).strip()
                    if "大一" in cell_str or "大二" in cell_str or "大三" in cell_str or "大四" in cell_str:
                        has_semester = True
                    if "模块" in cell_str or "选修课" in cell_str or "个性化" in cell_str:
                        has_module_info = True
        
        return has_semester and has_module_info
    
    def _parse_sydney_format(self, df, credit_df=None):
        """解析悉尼工商学院格式"""
        # 第一步：从学分食用概况表提取学分信息（如果有）
        credit_map = {}
        if credit_df is not None:
            credit_map = self._extract_credit_map(credit_df)
        
        # 第二步：解析课程与模块的对应关系
        course_module_map, standalone_modules = self._extract_course_module_relationships(df, credit_map)
        
        # 第三步：从关系映射中提取模块和课程
        modules, courses = self._build_modules_and_courses_from_map(course_module_map, credit_map, standalone_modules)
        
        # 第四步：组织层级结构
        self._map_sydney_to_hierarchy(modules, courses)
    
    def _extract_credit_map(self, credit_df):
        """从学分食用概况表提取课程学分映射"""
        credit_map = {}
        
        # 第一行是学分类型标题
        if len(credit_df) > 0:
            header_row = credit_df.iloc[0].values
            credit_types = []
            for j, cell in enumerate(header_row):
                if pd.notna(cell):
                    # 提取学分值，如"0.5学分" -> 0.5
                    cell_str = str(cell).strip()
                    match = re.search(r'(\d+(?:\.\d+)?)', cell_str)
                    if match:
                        credit_types.append((j, float(match.group(1))))
            
            # 从后续行提取课程
            for i in range(1, len(credit_df)):
                row = credit_df.iloc[i].values
                for j, credit_value in credit_types:
                    if j < len(row) and pd.notna(row[j]) and str(row[j]).strip():
                        course_name = str(row[j]).strip()
                        if course_name:
                            credit_map[course_name] = credit_value
        
        return credit_map
    
    def _extract_course_module_relationships(self, df, credit_map):
        """从Excel提取课程与模块的对应关系"""
        course_module_map = {}
        all_courses = []
        standalone_modules = []
        
        course_invalid_phrases = ["大一", "大二", "大三", "大四", "前半学期", "后半学期", "夏"]
        module_invalid_phrases = ["未必能完全", "注意", "推测", "简称建议", "选双学位", "各总计"]
        
        for i in range(len(df)):
            row = df.iloc[i].values
            
            # 提取这一行的课程（前5列）
            row_courses = []
            for j in range(5):
                if j < len(row) and pd.notna(row[j]) and str(row[j]).strip():
                    course_name = str(row[j]).strip()
                    
                    # 过滤无效内容
                    is_invalid = False
                    for phrase in course_invalid_phrases:
                        if phrase in course_name:
                            is_invalid = True
                            break
                    if is_invalid or len(course_name) < 2:
                        continue
                    
                    row_courses.append(course_name)
                    if course_name not in all_courses:
                        all_courses.append(course_name)
            
            # 提取这一行的模块（列5及以后）
            row_modules = []
            for j in range(5, len(row)):
                if pd.notna(row[j]) and str(row[j]).strip():
                    cell_content = str(row[j]).strip()
                    
                    # 清理模块名（先清理，再检查是否有效）
                    module_name = cell_content
                    
                    # 提取学分
                    credits = 0.0
                    credit_match = re.search(r'(\d+(?:\.\d+)?)\s*分', module_name)
                    if credit_match:
                        credits = float(credit_match.group(1))
                        # 从模块名中移除学分信息
                        module_name = re.sub(r'\s*\d+(?:\.\d+)?\s*分', '', module_name).strip()
                    
                    # 移除括号里的说明
                    module_name = re.sub(r'（[^）]*）', '', module_name).strip()
                    module_name = re.sub(r'\([^)]*\)', '', module_name).strip()
                    
                    # 检查清理后的模块名是否有效
                    is_invalid = False
                    for phrase in module_invalid_phrases:
                        if phrase in module_name:
                            is_invalid = True
                            break
                    if is_invalid:
                        continue
                    
                    # 判断是否是模块信息
                    if ("模块" in module_name or 
                        ("课" in module_name and "分" in module_name) or 
                        "必" in module_name or
                        "选修" in module_name or
                        "个性化" in module_name or
                        "创新创业" in module_name or
                        "劳动" in module_name or
                        "Canvas" in module_name or
                        "全英文" in module_name):
                        
                        if module_name:
                            row_modules.append((module_name, credits))
            
            # 将这一行的课程与模块建立关联
            if row_courses and row_modules:
                for course_name in row_courses:
                    if course_name not in course_module_map:
                        course_module_map[course_name] = []
                    for module_name, credits in row_modules:
                        if module_name not in [m[0] for m in course_module_map[course_name]]:
                            course_module_map[course_name].append((module_name, credits))
            elif row_modules:
                # 没有课程，但有模块 - 记录为独立模块
                for module_name, credits in row_modules:
                    if module_name not in standalone_modules:
                        standalone_modules.append((module_name, credits))
        
        # 处理没有对应模块的课程，将它们归到"专业课程"模块
        for course_name in all_courses:
            if course_name not in course_module_map:
                course_module_map[course_name] = [("专业课程", 0.0)]
        
        return course_module_map, standalone_modules
    
    def _parse_sydney_format_with_color(self, excel_path, sheet_name, credit_df=None):
        """解析悉尼工商学院格式（使用单元格颜色匹配）"""
        import openpyxl
        
        # 第一步：从学分食用概况表提取学分信息（如果有）
        credit_map = {}
        if credit_df is not None:
            credit_map = self._extract_credit_map(credit_df)
        
        # 第二步：使用openpyxl读取工作表，解析课程与模块的对应关系（基于颜色）
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb[sheet_name]
        course_module_map, standalone_modules = self._extract_course_module_relationships_by_color(sheet, credit_map)
        
        # 第三步：从关系映射中提取模块和课程
        modules, courses = self._build_modules_and_courses_from_map(course_module_map, credit_map, standalone_modules)
        
        # 第四步：组织层级结构
        self._map_sydney_to_hierarchy(modules, courses)
    
    def _extract_course_module_relationships_by_color(self, sheet, credit_map):
        """从Excel提取课程与模块的对应关系（基于同一行，颜色作为补充）"""
        course_module_map = {}
        all_courses = []
        standalone_modules = []
        
        course_invalid_phrases = ["大一", "大二", "大三", "大四", "前半学期", "后半学期", "夏"]
        module_invalid_phrases = ["未必能完全", "注意", "推测", "简称建议", "选双学位", "各总计"]
        
        # 第一次遍历：收集课程和同一行的模块
        for row_idx in range(1, sheet.max_row + 1):
            row = sheet[row_idx]
            
            # 提取这一行的课程（前5列）
            row_courses = []
            for col_idx in range(1, 6):
                if col_idx - 1 >= len(row):
                    continue
                cell = row[col_idx - 1]
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                
                course_name = str(cell.value).strip()
                
                # 过滤无效内容
                is_invalid = False
                for phrase in course_invalid_phrases:
                    if phrase in course_name:
                        is_invalid = True
                        break
                if is_invalid or len(course_name) < 2:
                    continue
                
                row_courses.append(course_name)
                if course_name not in all_courses:
                    all_courses.append(course_name)
            
            # 提取这一行的模块（列6及以后）
            row_modules = []
            for col_idx in range(6, sheet.max_column + 1):
                if col_idx - 1 >= len(row):
                    continue
                cell = row[col_idx - 1]
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                
                cell_content = str(cell.value).strip()
                
                # 清理模块名
                module_name = cell_content
                
                # 提取学分
                credits = 0.0
                credit_match = re.search(r'(\d+(?:\.\d+)?)\s*分', module_name)
                if credit_match:
                    credits = float(credit_match.group(1))
                    module_name = re.sub(r'\s*\d+(?:\.\d+)?\s*分', '', module_name).strip()
                
                # 移除括号里的说明
                module_name = re.sub(r'（[^）]*）', '', module_name).strip()
                module_name = re.sub(r'\([^)]*\)', '', module_name).strip()
                
                # 检查清理后的模块名是否有效
                is_invalid = False
                for phrase in module_invalid_phrases:
                    if phrase in module_name:
                        is_invalid = True
                        break
                if is_invalid:
                    continue
                
                # 判断是否是模块信息
                if ("模块" in module_name or 
                    ("课" in module_name and "分" in module_name) or 
                    "必" in module_name or
                    "选修" in module_name or
                    "个性化" in module_name or
                    "创新创业" in module_name or
                    "劳动" in module_name or
                    "Canvas" in module_name or
                    "全英文" in module_name):
                    
                    if module_name:
                        row_modules.append((module_name, credits))
            
            # 将这一行的课程与模块建立关联
            if row_courses and row_modules:
                for course_name in row_courses:
                    if course_name not in course_module_map:
                        course_module_map[course_name] = []
                    for module_name, credits in row_modules:
                        if module_name not in [m[0] for m in course_module_map[course_name]]:
                            course_module_map[course_name].append((module_name, credits))
            elif row_modules:
                # 没有课程，但有模块 - 记录为独立模块
                for module_name, credits in row_modules:
                    if module_name not in standalone_modules:
                        standalone_modules.append((module_name, credits))
        
        # 处理没有对应模块的课程，将它们归到"专业课程"模块
        for course_name in all_courses:
            if course_name not in course_module_map:
                course_module_map[course_name] = [("专业课程", 0.0)]
        
        return course_module_map, standalone_modules
    
    def _build_modules_and_courses_from_map(self, course_module_map, credit_map, standalone_modules=None):
        """从课程-模块映射中构建模块和课程列表"""
        # 先收集所有模块
        module_dict = {}
        for course_name, course_modules in course_module_map.items():
            for module_name, credits in course_modules:
                if module_name not in module_dict:
                    module_dict[module_name] = {
                        'name': module_name,
                        'credits': credits,
                        'order': len(module_dict) + 1
                    }
        
        # 添加独立模块
        if standalone_modules:
            for module_name, credits in standalone_modules:
                if module_name not in module_dict:
                    module_dict[module_name] = {
                        'name': module_name,
                        'credits': credits,
                        'order': len(module_dict) + 1
                    }
        
        # 转换为模块列表
        result_modules = []
        for module_name, module_info in module_dict.items():
            result_modules.append({
                'order': module_info['order'],
                'sub_order': None,
                'name': module_info['name'],
                'credits': module_info['credits'],
                'parent_order': None
            })
        
        # 创建模块名到ID的映射
        module_name_to_id = {m['name']: m['order'] for m in result_modules}
        
        # 构建课程列表
        result_courses = []
        course_id = 100
        for course_name, course_modules in course_module_map.items():
            # 查找学分
            credits = credit_map.get(course_name, 0.0)
            
            # 获取所有关联的模块ID
            module_ids = []
            for module_name, _ in course_modules:
                if module_name in module_name_to_id:
                    module_ids.append(module_name_to_id[module_name])
            
            # 课程可以对应多个模块，但我们需要选择一个主要模块（这里选择第一个）
            primary_module_id = module_ids[0] if module_ids else 1
            
            result_courses.append({
                'id': course_id,
                'course_code': f'SYDNEY{course_id}',
                'name': course_name,
                'credits': credits,
                'module_id': primary_module_id,
                'module_ids': module_ids  # 保存所有模块ID
            })
            course_id += 1
        
        return result_modules, result_courses
    
    def _map_sydney_to_hierarchy(self, modules, courses):
        """将悉尼格式解析结果映射为层级结构"""
        # 组织模块
        hierarchy_modules = []
        module_id_map = {}
        current_id = 1
        
        for mod in modules:
            hierarchy_modules.append({
                'id': current_id,
                'name': mod['name'],
                'required_credits': mod['credits'],
                'parent_id': mod.get('parent_order'),
                'source': 'excel'
            })
            module_id_map[mod['order']] = current_id
            current_id += 1
        
        # 创建模块ID到名称的映射
        module_id_to_name = {mod['id']: mod['name'] for mod in hierarchy_modules}
        
        # 更新课程，映射module_ids并添加module_names
        for course in courses:
            # 映射module_ids到新的ID（如果需要的话）
            if 'module_ids' in course:
                new_module_ids = []
                for old_id in course['module_ids']:
                    if old_id in module_id_map:
                        new_module_ids.append(module_id_map[old_id])
                course['module_ids'] = new_module_ids
                
                # 添加module_names字段，用于前端显示
                course['module_names'] = [module_id_to_name[mid] for mid in new_module_ids if mid in module_id_to_name]
            
            # 同时也更新module_id
            if 'module_id' in course and course['module_id'] in module_id_map:
                course['module_id'] = module_id_map[course['module_id']]
        
        self.modules = hierarchy_modules
        self.courses = courses

    def _extract_major_from_sheet_name(self, sheet_name):
        """从工作表名提取专业"""
        clean_name = sheet_name.replace('(直招)', '').replace('直招', '').strip()
        for major, college in MAJOR_COLLEGE_MAP.items():
            if clean_name in major or major in clean_name:
                self.major_name = major
                self.college_name = college
                break

    def _parse_modules_from_excel(self, df):
        """从Excel解析模块，按照数字编号构建层级结构"""
        modules = []
        current_root_order = 0
        special_modules_order = {}  # 保存特殊模块的原始顺序

        for idx, row in df.iterrows():
            cell0 = str(row[0]).strip() if pd.notna(row[0]) else ""

            # 匹配根模块：1. 模块名 或 1、模块名
            root_match = re.match(r'^(\d+)[\.、]\s*([^\n]+?)(?:\s*要求.*)?$', cell0)
            if root_match:
                order = int(root_match.group(1))
                name = root_match.group(2).strip()
                
                # 跳过"专业教育模块"，因为它的子模块会被提升为主模块
                if name == "专业教育模块":
                    current_root_order = order  # 保存当前编号，用于子模块
                    continue
                
                current_root_order = order

                # 查找学分
                credits = self._find_module_credits(df, idx)

                modules.append({
                    'order': order,
                    'sub_order': None,
                    'name': name,
                    'credits': credits,
                    'parent_order': None
                })
                continue

            # 匹配子模块：（1）模块名 或 (1) 模块名
            sub_match = re.match(r'^[（(](\d+)[）)]\s*([^\n]+?)(?:\s*要求.*)?$', cell0)
            if sub_match and current_root_order > 0:
                sub_order = int(sub_match.group(1))
                name = sub_match.group(2).strip()

                # 查找学分
                credits = self._find_module_credits(df, idx)

                # 检查是否是专业基础课程、专业必修课程、专业选修课程，如果是则提升为主模块
                if name in ["专业基础课程", "专业必修课程", "专业选修课程"]:
                    # 保存原始顺序位置：current_root_order + sub_order * 0.1
                    sort_key = current_root_order + sub_order * 0.1
                    
                    modules.append({
                        'order': sort_key,  # 使用带小数的排序键
                        'sub_order': None,
                        'name': name,
                        'credits': credits,
                        'parent_order': None
                    })
                else:
                    # 普通子模块
                    modules.append({
                        'order': current_root_order,
                        'sub_order': sub_order,
                        'name': name,
                        'credits': credits,
                        'parent_order': current_root_order
                    })
                continue

        # 去重和排序：使用更严格的key
        seen = set()
        unique_modules = []
        for mod in modules:
            # 使用更完整的key来去重，包括order和sub_order
            key = (mod['order'], mod.get('sub_order'), mod['name'], mod['credits'])
            if key not in seen:
                seen.add(key)
                unique_modules.append(mod)

        # 在排序之前，先保存原始的order到临时字段
        for mod in unique_modules:
            mod['original_order'] = mod['order']
        
        # 排序：先按order（包括带小数的排序键），再处理子模块
        unique_modules.sort(key=lambda x: (x['order'], x.get('sub_order') is None, x.get('sub_order') or 0))
        
        # 重新分配整数order，保持顺序，并更新子模块的parent_order
        order_map = {}  # 旧order -> 新order
        current_order = 1
        for mod in unique_modules:
            if mod['sub_order'] is None:
                old_order = mod['original_order']  # 使用原始order
                mod['order'] = current_order
                order_map[old_order] = current_order
                current_order += 1
        
        # 更新子模块的parent_order
        for mod in unique_modules:
            if mod['sub_order'] is not None:
                old_parent = mod['parent_order']
                if old_parent in order_map:
                    mod['parent_order'] = order_map[old_parent]
                    mod['order'] = order_map[mod['original_order']]  # 使用原始order
        
        # 清理临时字段
        for mod in unique_modules:
            if 'original_order' in mod:
                del mod['original_order']
        
        self._map_to_hierarchy_structure(unique_modules)

    def _find_module_credits(self, df, start_idx):
        """在模块行附近查找学分信息"""
        for i in range(start_idx, min(start_idx + 5, len(df))):
            for col_idx in range(len(df.columns)):
                cell = str(df.iloc[i, col_idx]).strip() if pd.notna(df.iloc[i, col_idx]) else ""

                # 匹配：要求最低学分：XX 或 要求：XX学分
                credit_match = re.search(r'(?:要求最低学分|要求)[:：]\s*(\d+(?:\.\d+)?)', cell)
                if credit_match:
                    return float(credit_match.group(1))

                # 匹配：XX学分
                credit_match2 = re.search(r'(\d+(?:\.\d+)?)\s*学分', cell)
                if credit_match2:
                    return float(credit_match2.group(1))

        # 如果没找到，返回0.0（不设置默认值）
        return 0.0

    def _map_to_hierarchy_structure(self, parsed_modules):
        """将解析的模块按照层级结构组织"""
        modules = []
        module_id = 1
        order_id_map = {}

        # 先创建所有根模块
        root_modules = [m for m in parsed_modules if m['sub_order'] is None]
        for mod in root_modules:
            modules.append({
                'id': module_id,
                'name': mod['name'],
                'required_credits': mod['credits'],
                'parent_id': None,
                'source': 'excel'
            })
            order_id_map[mod['order']] = module_id
            module_id += 1

        # 然后创建所有子模块
        sub_modules = [m for m in parsed_modules if m['sub_order'] is not None]
        for mod in sub_modules:
            parent_id = order_id_map.get(mod['parent_order'])
            if parent_id:
                modules.append({
                    'id': module_id,
                    'name': mod['name'],
                    'required_credits': mod['credits'],
                    'parent_id': parent_id,
                    'source': 'excel'
                })
                module_id += 1

        self.modules = modules

    def _parse_courses_from_excel(self, df):
        """从Excel解析课程"""
        courses = []
        course_id = 100
        
        # 创建模块名到ID的映射
        module_name_to_id = {}
        for mod in self.modules:
            module_name_to_id[mod['name']] = mod['id']
        
        # 识别表头行（包含"课程代码"的行）
        header_idx = -1
        for idx, row in df.iterrows():
            row_str = str(row).lower()
            if '课程代码' in row_str or 'course code' in row_str:
                header_idx = idx
                break
        
        # 如果没找到表头，尝试从第3行开始
        start_idx = header_idx + 1 if header_idx >= 0 else 3
        
        for idx in range(start_idx, len(df)):
            row = df.iloc[idx]
            
            # 列0：课程分类
            col0 = str(row[0]).strip() if pd.notna(row[0]) else ""
            # 列1：课程代码
            col1 = str(row[1]).strip() if pd.notna(row[1]) else ""
            # 列2：课程名称
            col2 = str(row[2]).strip() if pd.notna(row[2]) else ""
            # 列3：课程英文名称
            col3 = str(row[3]).strip() if pd.notna(row[3]) else ""
            # 列4：学分
            col4 = row[4] if pd.notna(row[4]) else ""
            
            # 跳过非课程行（如合计、空行、子模块行等）
            if '合计' in col0 or '合计' in col2:
                continue
            if not col1 and not col2:
                continue
            if '模块' in col0 or re.match(r'^[（(]\d+[）)]', col0):
                continue
            
            # 尝试从列1提取课程代码（如：GBK2000001）
            code_match = re.match(r'([A-Z]{2,4}\d{4,8})', col1)
            if code_match:
                course_code = code_match.group(1)
            else:
                # 如果列1不是课程代码，可能在列0或列2
                code_match = re.search(r'([A-Z]{2,4}\d{4,8})', col0 + ' ' + col2)
                if code_match:
                    course_code = code_match.group(1)
                elif col1 and len(col1) <= 20 and col1 != '详见方案':
                    # 使用列1作为课程代码
                    course_code = col1
                else:
                    continue
            
            # 课程名称
            course_name = col2 if col2 else col3
            if not course_name or '详见方案' in course_name:
                continue
            
            # 学分
            try:
                credits = float(col4)
            except (ValueError, TypeError):
                credits = 0.0
            
            # 确定所属模块（根据课程分类或位置）
            module_ids = []
            module_names = []
            
            # 从课程分类中匹配模块
            for module_name, module_id in module_name_to_id.items():
                if col0 and (module_name in col0 or col0 in module_name):
                    if module_id not in module_ids:
                        module_ids.append(module_id)
                        module_names.append(module_name)
            
            # 如果没有匹配到，使用单个模块
            if not module_ids:
                primary_module_id = self._find_course_module_id(df, idx)
                module_ids = [primary_module_id]
                # 查找模块名称
                for module_name, module_id in module_name_to_id.items():
                    if module_id == primary_module_id:
                        module_names = [module_name]
                        break
            
            primary_module_id = module_ids[0] if module_ids else 1
            
            courses.append({
                'id': course_id,
                'course_code': course_code,
                'name': course_name,
                'credits': credits,
                'module_id': primary_module_id,
                'module_ids': module_ids,
                'module_names': module_names
            })
            course_id += 1
        
        self.courses = courses

    def _find_course_module_id(self, df, course_idx):
        """根据课程位置确定所属模块"""
        # 简单实现：返回第一个根模块的ID，或者让用户在预览时手动设置
        # 这里返回None，用户可以在预览页面调整
        return 1

    def get_result(self):
        """获取解析结果"""
        return {
            'college_name': self.college_name,
            'major_name': self.major_name,
            'filename': self.filename,
            'modules': self.modules,
            'courses': self.courses
        }


# ==================== Flask路由 ====================
@app.route('/', methods=['GET', 'POST'])
def index():
    """首页 - 上传PDF或Excel"""
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('index.html', error='请选择文件')
        
        file = request.files['file']
        if file.filename == '':
            return render_template('index.html', error='请选择文件')
        
        # 处理PDF文件
        if file and file.filename.lower().endswith('.pdf'):
            filename = str(uuid.uuid4()) + '.pdf'
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # 解析PDF
            parser = SmartPDFParser()
            try:
                parser.parse(file_path, file.filename)
                result = parser.get_result()
                
                session_id = os.path.splitext(filename)[0]
                parsing_results[session_id] = result
                
                return redirect(url_for('preview', session_id=session_id))
            except Exception as e:
                return render_template('index.html', error=f'解析失败: {str(e)}')
        
        # 处理Excel文件
        elif file and file.filename.lower().endswith(('.xlsx', '.xls')):
            filename = str(uuid.uuid4()) + os.path.splitext(file.filename)[1]
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # 获取Excel工作表列表
            try:
                xl = pd.ExcelFile(file_path)
                sheet_names = xl.sheet_names
                
                # 保存文件路径，供选择工作表时使用
                session_id = os.path.splitext(filename)[0]
                parsing_results[session_id] = {
                    'file_path': file_path,
                    'filename': file.filename,
                    'sheet_names': sheet_names
                }
                
                return redirect(url_for('select_sheet', session_id=session_id))
            except Exception as e:
                return render_template('index.html', error=f'读取Excel失败: {str(e)}')
        
        else:
            return render_template('index.html', error='请上传PDF或Excel文件')
    
    # 获取所有学院列表供选择
    colleges = get_all_colleges()
    return render_template('index.html', colleges=colleges)


@app.route('/select_sheet/<session_id>', methods=['GET', 'POST'])
def select_sheet(session_id):
    """选择Excel工作表"""
    if session_id not in parsing_results:
        return redirect(url_for('index'))
    
    data = parsing_results[session_id]
    
    if request.method == 'POST':
        sheet_name = request.form.get('sheet_name', '')
        if not sheet_name:
            return render_template('select_sheet.html', 
                                   session_id=session_id,
                                   sheet_names=data.get('sheet_names', []),
                                   error='请选择工作表')
        
        # 解析Excel
        parser = SmartExcelParser()
        try:
            parser.parse(data['file_path'], sheet_name, data['filename'])
            result = parser.get_result()
            
            parsing_results[session_id] = result
            
            return redirect(url_for('preview', session_id=session_id))
        except Exception as e:
            return render_template('select_sheet.html', 
                                   session_id=session_id,
                                   sheet_names=data.get('sheet_names', []),
                                   error=f'解析失败: {str(e)}')
    
    return render_template('select_sheet.html', 
                           session_id=session_id,
                           sheet_names=data.get('sheet_names', []))


@app.route('/preview/<session_id>')
def preview(session_id):
    """预览页面"""
    if session_id not in parsing_results:
        return redirect(url_for('index'))
    
    result = parsing_results[session_id]
    
    # 获取所有学院和对应专业供选择
    colleges = get_all_colleges()
    majors = get_majors_by_college(result.get('college_name', ''))
    
    return render_template('preview.html', 
                           data=result, 
                           session_id=session_id,
                           colleges=colleges,
                           majors=majors)


@app.route('/get_majors/<college_name>')
def get_majors(college_name):
    """根据学院获取专业列表（AJAX接口）"""
    majors = get_majors_by_college(college_name)
    return jsonify({'success': True, 'majors': majors})


@app.route('/update', methods=['POST'])
def update():
    """更新解析结果"""
    session_id = request.form.get('session_id')
    if session_id not in parsing_results:
        return jsonify({'success': False, 'message': '会话不存在'})
    
    try:
        data = parsing_results[session_id]
        data['college_name'] = request.form.get('college_name', '')
        data['major_name'] = request.form.get('major_name', '')
        
        modules = []
        for key in request.form:
            if key.startswith('module_name_'):
                idx = int(key.split('_')[-1])
                modules.append({
                    'id': int(request.form.get(f'module_id_{idx}', idx)),
                    'name': request.form.get(f'module_name_{idx}', ''),
                    'required_credits': float(request.form.get(f'module_credits_{idx}', 0)),
                    'parent_id': int(request.form.get(f'module_parent_{idx}', 0)) or None
                })
        
        if modules:
            data['modules'] = modules
        
        courses = []
        for key in request.form:
            if key.startswith('course_code_'):
                idx = int(key.split('_')[-1])
                courses.append({
                    'id': int(request.form.get(f'course_id_{idx}', idx)),
                    'course_code': request.form.get(f'course_code_{idx}', ''),
                    'name': request.form.get(f'course_name_{idx}', ''),
                    'credits': float(request.form.get(f'course_credit_{idx}', 0)),
                    'module_id': int(request.form.get(f'course_module_{idx}', 0))
                })
        
        if courses:
            data['courses'] = courses
        
        parsing_results[session_id] = data
        return jsonify({'success': True})
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/test_db')
def test_db():
    """测试数据库连接"""
    try:
        conn = pymysql.connect(**DB_CONFIG)
        cursor = conn.cursor()
        
        # 检查必要的表是否存在
        tables = ['college', 'major', 'module', 'course']
        missing_tables = []
        
        for table in tables:
            cursor.execute(f"SHOW TABLES LIKE '{table}'")
            if not cursor.fetchone():
                missing_tables.append(table)
        
        cursor.close()
        conn.close()
        
        if missing_tables:
            return jsonify({
                'success': False,
                'message': f'缺少必要的表: {", ".join(missing_tables)}',
                'tables': {'found': [t for t in tables if t not in missing_tables], 'missing': missing_tables}
            })
        
        return jsonify({
            'success': True,
            'message': '数据库连接正常，所有表都存在',
            'tables': tables
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'数据库连接失败: {str(e)}',
            'config': {k: v if k != 'password' else '***' for k, v in DB_CONFIG.items()}
        })


@app.route('/import/<session_id>')
def import_db(session_id):
    """导入数据库"""
    if session_id not in parsing_results:
        return jsonify({'success': False, 'message': '会话不存在'})
    
    data = parsing_results[session_id]
    
    try:
        print(f"🔄 正在导入数据库...")
        print(f"   学院: {data['college_name']}")
        print(f"   专业: {data['major_name']}")
        
        # 连接数据库
        conn = pymysql.connect(**DB_CONFIG)
        cursor = conn.cursor()
        print("✅ 数据库连接成功")
        
        # 处理学院
        cursor.execute("SELECT id FROM college WHERE name = %s", (data['college_name'],))
        college = cursor.fetchone()
        if not college:
            cursor.execute("INSERT INTO college (name) VALUES (%s)", 
                          (data['college_name'],))
            conn.commit()
            college_id = cursor.lastrowid
            print(f"✅ 创建学院: {data['college_name']} (ID: {college_id})")
        else:
            college_id = college[0]
            print(f"✅ 学院已存在: {data['college_name']} (ID: {college_id})")
        
        # 处理专业
        cursor.execute("SELECT id FROM major WHERE college_id = %s AND name = %s", 
                      (college_id, data['major_name']))
        major = cursor.fetchone()
        if not major:
            cursor.execute("INSERT INTO major (name, college_id) VALUES (%s, %s)",
                          (data['major_name'], college_id))
            conn.commit()
            major_id = cursor.lastrowid
            print(f"✅ 创建专业: {data['major_name']} (ID: {major_id})")
        else:
            major_id = major[0]
            print(f"✅ 专业已存在: {data['major_name']} (ID: {major_id})")
        
        # 删除旧数据（先删除子模块，再删除父模块，处理自引用外键）
        cursor.execute("SELECT id FROM module WHERE major_id = %s", (major_id,))
        old_mods = [row[0] for row in cursor.fetchall()]
        if old_mods:
            cursor.execute(f"DELETE FROM course WHERE module_id IN ({','.join(str(m) for m in old_mods)})")
            # 循环删除，直到所有模块都被删除（先删除叶子节点）
            deleted_count = 0
            while old_mods:
                # 使用临时表方式避免 MySQL 子查询更新同表的限制
                cursor.execute("""
                    DELETE m FROM module m
                    LEFT JOIN module p ON m.id = p.parent_id
                    WHERE m.major_id = %s AND p.parent_id IS NULL
                """, (major_id,))
                deleted = cursor.rowcount
                if deleted == 0:
                    break
                deleted_count += deleted
                conn.commit()
                cursor.execute("SELECT id FROM module WHERE major_id = %s", (major_id,))
                old_mods = [row[0] for row in cursor.fetchall()]
            print(f"✅ 删除旧数据: {deleted_count} 个模块")
        
        # 插入模块（使用数据库自动生成ID）
        mod_count = 0
        old_new_id_map = {}  # 旧ID -> 新ID映射
        
        # 先插入所有模块（不指定ID）
        for mod in data['modules']:
            try:
                cursor.execute("""
                    INSERT INTO module (name, required_credits, parent_id, college_id, major_id)
                    VALUES (%s, %s, %s, %s, %s)
                """, (mod['name'], mod['required_credits'], None, college_id, major_id))
                new_id = cursor.lastrowid
                old_new_id_map[mod['id']] = new_id
                mod_count += 1
            except Exception as e:
                print(f"⚠️ 模块 {mod['name']} 插入失败: {e}")
        conn.commit()
        print(f"✅ 插入模块: {mod_count} 个")
        
        # 更新父模块ID（使用新的ID）
        for mod in data['modules']:
            if mod['parent_id'] and mod['parent_id'] in old_new_id_map:
                new_parent_id = old_new_id_map[mod['parent_id']]
                new_id = old_new_id_map[mod['id']]
                cursor.execute("UPDATE module SET parent_id = %s WHERE id = %s", (new_parent_id, new_id))
        conn.commit()
        print(f"✅ 更新父模块关系")
        
        # 插入课程
        new_count = 0
        skip_count = 0
        for course in data['courses']:
            cursor.execute("SELECT id FROM course WHERE course_code = %s", (course['course_code'],))
            if not cursor.fetchone():
                try:
                    # 将课程的模块ID映射到新的ID
                    module_id = old_new_id_map.get(course['module_id'], course['module_id'])
                    cursor.execute("""
                        INSERT INTO course (course_code, name, credits, module_id)
                        VALUES (%s, %s, %s, %s)
                    """, (course['course_code'], course['name'], course['credits'], module_id))
                    new_count += 1
                except Exception as e:
                    print(f"⚠️ 课程 {course['course_code']} 插入失败: {e}")
            else:
                skip_count += 1
        conn.commit()
        print(f"✅ 插入课程: {new_count} 个 (跳过重复: {skip_count} 个)")
        
        cursor.close()
        conn.close()
        print("✅ 数据库连接已关闭")
        
        # 清除主应用的缓存
        try:
            import requests
            response = requests.post('http://localhost:5000/api/cache/clear')
            if response.json().get('success'):
                print("✅ 已清除主应用缓存")
            else:
                print("⚠️ 清除缓存失败")
        except Exception as e:
            print(f"⚠️ 清除缓存失败: {e}")
        
        del parsing_results[session_id]
        
        return jsonify({
            'success': True,
            'message': f'导入成功！模块: {len(data["modules"])}, 课程: {new_count}'
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/preview/<session_id>')
def api_preview(session_id):
    """获取预览数据"""
    if session_id not in parsing_results:
        return jsonify({'success': False, 'message': '会话不存在'})
    
    return jsonify({'success': True, 'data': parsing_results[session_id]})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
