"""分析悉尼工商学院Excel的单元格格式特征"""
import os
import openpyxl

excel_path = os.path.join(os.path.dirname(__file__), 'shucredit_scripts', 'pdf_app', 'uploads', 'sydney_business_25.xlsx')

print("="*100)
print("分析悉尼工商学院Excel单元格格式特征")
print("="*100)

wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb['信管课程25级']

# 分析前50行的单元格格式
print("\n分析前50行的单元格格式:")
print("-"*100)

for row_idx in range(1, min(51, sheet.max_row + 1)):
    row = sheet[row_idx]
    has_data = False
    row_info = []
    
    for col_idx, cell in enumerate(row, 1):
        if cell.value is None or str(cell.value).strip() == "":
            continue
        
        has_data = True
        cell_info = []
        
        # 列号和值
        cell_info.append(f"[{col_idx}] {cell.value}")
        
        # 字体样式
        font_info = []
        if cell.font:
            if cell.font.name:
                font_info.append(f"字体:{cell.font.name}")
            if cell.font.size:
                font_info.append(f"字号:{cell.font.size}")
            if cell.font.bold:
                font_info.append("加粗")
            if cell.font.italic:
                font_info.append("斜体")
            if cell.font.color and cell.font.color.rgb:
                font_info.append(f"颜色:{cell.font.color.rgb}")
        
        # 边框样式
        border_info = []
        if cell.border:
            for side in ['left', 'right', 'top', 'bottom']:
                if hasattr(cell.border, side):
                    side_border = getattr(cell.border, side)
                    if side_border and side_border.style:
                        border_info.append(f"{side}:{side_border.style}")
        
        # 填充样式
        fill_info = []
        if cell.fill:
            if hasattr(cell.fill, 'patternType') and cell.fill.patternType:
                fill_info.append(f"图案:{cell.fill.patternType}")
            if hasattr(cell.fill, 'fgColor') and cell.fill.fgColor and cell.fill.fgColor.rgb:
                fill_info.append(f"前景:{cell.fill.fgColor.rgb}")
            if hasattr(cell.fill, 'bgColor') and cell.fill.bgColor and cell.fill.bgColor.rgb:
                fill_info.append(f"背景:{cell.fill.bgColor.rgb}")
        
        if font_info:
            cell_info.append(f"({', '.join(font_info)})")
        if border_info:
            cell_info.append(f"[{', '.join(border_info)}]")
        if fill_info:
            cell_info.append(f"<{', '.join(fill_info)}>")
        
        row_info.append(" ".join(cell_info))
    
    if has_data:
        print(f"\n行{row_idx}:")
        for info in row_info:
            print(f"  {info}")

print("\n" + "="*100)
print("尝试按填充颜色分组分析:")
print("="*100)

# 收集所有课程单元格，按填充颜色分组
color_groups = {}
for row_idx in range(1, sheet.max_row + 1):
    row = sheet[row_idx]
    for col_idx in range(1, min(6, sheet.max_column + 1)):  # 前5列是课程列
        cell = row[col_idx - 1]
        if cell.value is None or str(cell.value).strip() == "":
            continue
        
        cell_value = str(cell.value).strip()
        # 过滤无效值
        invalid_phrases = ["大一", "大二", "大三", "大四", "前半学期", "后半学期", "夏"]
        is_invalid = any(phrase in cell_value for phrase in invalid_phrases)
        if is_invalid or len(cell_value) < 2:
            continue
        
        # 获取填充颜色
        color_key = "无填充"
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            color_key = cell.fill.fgColor.rgb
        
        if color_key not in color_groups:
            color_groups[color_key] = []
        color_groups[color_key].append(cell_value)

for color, courses in color_groups.items():
    print(f"\n颜色: {color}")
    print(f"  课程数: {len(courses)}")
    print(f"  课程: {', '.join(courses[:10])}")
    if len(courses) > 10:
        print(f"  ... 还有 {len(courses) - 10} 门")
